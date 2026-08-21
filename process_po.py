#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
process_po.py (v2) — Tự động điền PO (PDF) từ nhiều hệ thống siêu thị vào form BÁN (Form_Hàng_MT10.xlsx)

Hỗ trợ 12 hệ thống: MiniStop, WinMart, WinMart+, CircleK, Co-opFood (JDA), Co-opSmile,
Co-opCheers, Sài Gòn HD, OsiFood, Satrafood, 3Sach, Siba.

CÁCH DÙNG:
  1. Bỏ vào 1 thư mục: tất cả PDF PO của ngày hôm đó + file Form_Hàng_MT10.xlsx mới nhất.
  2. Chạy:  python3 process_po.py "/đường/dẫn/tới/thư_mục"
  3. Kết quả: "Form_Hàng_MT10_đã_điền.xlsx" được tạo trong CHÍNH thư mục đó.

YÊU CẦU: pip install pdfplumber openpyxl --break-system-packages
"""

import sys
import os
import re
import glob
import shutil
import unicodedata
import datetime
import difflib

import pdfplumber
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

# =============================================================================
# 1. BẢNG MAPPING BARCODE -> MÃ SẢN PHẨM NỘI BỘ (TP0xxxxx)
# =============================================================================

BARCODE_TO_TP = {
    "8935117700098": "TP010004",  # Flan 100gr lốc 2      -> FĐ2 100g
    "8935117700104": "TP010005",  # Flan 100gr lốc 6       -> FĐ6 100g
    "8935117700012": "TP010008",  # Flan 54gr lốc 10       -> FL10 54g
    "8935117700036": "TP010007",  # Flan 54gr lốc 6        -> FL6 54g
    "8935117700128": "TP010014",  # Flan Caramel cao cấp 80gr (Lốc) -> FCC Lốc
    "8935117700135": "TP010016",  # Flan Caramel CFSD 82gr (Lốc)    -> CFSD Lốc
    "8935117702016": "TP050002",  # Thạch dừa nhỏ 90gr             -> TDN 90g
    "8935117702023": "TP050005",  # Thạch dừa lớn 190gr (Lốc)      -> TDL 190g
    "8935117702115": "TP050004",  # Thạch dừa lớn 190gr (Ly)       -> TDL 190g LY
    "8935117702221": "TP050037",  # Thạch dừa vị đào (Lốc)         -> ĐÀO
    "8935117702177": "TP050038",  # Thạch dừa vị đào (Ly)          -> ĐÀO Ly
    "8935117702306": "TP050039",  # Thạch dừa dưa lưới (Lốc)       -> DƯA LƯỚI
    "8935117702252": "TP050040",  # Thạch dừa dưa lưới (Ly)        -> DƯA LƯỚI Ly
    "8935117702030": "TP050007",  # Thạch dừa bịch 400gr           -> TDB 400g
    "8935117702047": "TP050006",  # Thạch dừa bịch 900gr           -> TDB 900g
    "8935117702368": "TP050061",  # Cocochew chanh dây 300gr       -> Jelly Chanh Dây
    "8935117702375": "TP050062",  # Cocochew dứa 300gr             -> Jelly Dứa
    "8935117703037": "TP030039",  # Yuzu túi 500gr                 -> XÁ 1/2kg
    "8935117703044": "TP030040",  # Yuzu túi 1kg                   -> XÁ 1kg
    "8935117703068": "TP030019",  # Yuzu thùng 6kg                 -> XÁ 6kg
    "8935117721024": "TP020061",  # STC Yuzu dâu                   -> DÂU
    "8935117721031": "TP020059",  # STC Yuzu nho                   -> NHO
    "8935117721017": "TP020057",  # STC Yuzu táo                   -> TÁO
    "8935117721000": "TP020055",  # STC Yuzu cam                   -> CAM
}

HOP_VARIANT = {
    "8935117700128": "TP010013",  # FCC Hộp
    "8935117700135": "TP010017",  # CFSD Hộp
}

TP_TO_COL = {
    "TP010004": "G", "TP010005": "H", "TP010007": "I", "TP010008": "J",
    "TP010014": "K", "TP010013": "L", "TP010016": "M", "TP010017": "N",
    "TP050002": "O", "TP050005": "P", "TP050004": "Q", "TP050006": "R", "TP050007": "S",
    "TP030039": "T", "TP030040": "U", "TP030019": "V",
    "TP050061": "W", "TP050062": "X",
    "TP020055": "Y", "TP020057": "Z", "TP020059": "AA", "TP020061": "AB",
    "TP050039": "AC", "TP050040": "AD", "TP050037": "AE", "TP050038": "AF",
}

# --- Quy tắc riêng của MiniStop ---
MINISTOP_TDL_LY_BARCODE = "8935117702115"
MINISTOP_TDL_LOC_TP = "TP050005"
UNITS_PER_LOC = 6

# --- Quy tắc riêng của 3Sach: PO đặt theo đơn vị "Ly" nhưng khi điền vào form
# luôn tính vào cột "Lốc" tương ứng, số lượng chia 6 (1 Lốc = 6 Ly). ---
SACH3_LY_TO_LOC_TP = {
    "8935117702115": "TP050005",  # Thạch dừa lớn 190g: Ly -> Lốc
    "8935117702177": "TP050037",  # Thạch dừa vị đào: Ly -> Lốc
    "8935117702252": "TP050039",  # Thạch dừa dưa lưới: Ly -> Lốc
}


def resolve_tp(barcode, uom_hint=None, vendor=None, qty=None):
    if vendor == "MINISTOP" and barcode == MINISTOP_TDL_LY_BARCODE:
        return MINISTOP_TDL_LOC_TP, round((qty or 0) / UNITS_PER_LOC)
    if vendor == "3SACH" and barcode in SACH3_LY_TO_LOC_TP:
        return SACH3_LY_TO_LOC_TP[barcode], round((qty or 0) / UNITS_PER_LOC)
    if uom_hint and uom_hint.strip().lower() in ("hop", "hộp") and barcode in HOP_VARIANT:
        return HOP_VARIANT[barcode], qty
    return BARCODE_TO_TP.get(barcode), qty


def _normalize(s):
    """Bỏ dấu tiếng Việt, viết hoa, gộp khoảng trắng — dùng để so khớp mờ tên/địa chỉ."""
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode()
    s = re.sub(r'[^A-Za-z0-9\s]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip().upper()


# =============================================================================
# 2. PARSER CHO TỪNG HỆ THỐNG PDF
# =============================================================================

def parse_ministop(path):
    orders = []
    fname = os.path.basename(path)
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            if "MINISTOP" not in text.upper():
                continue
            m = re.search(r'^(\d{3,6})-\d{10,}\s', text, re.M)
            if not m:
                print(f"  [CẢNH BÁO] {fname} trang {page.page_number}: nhận diện MINISTOP "
                      f"nhưng không đọc được mã cửa hàng -> BỎ QUA, cần kiểm tra tay")
                continue
            store_code = m.group(1)
            po_match = re.search(r'^(\d{3,6}-\d{10,})\s', text, re.M)
            po_ref = po_match.group(1) if po_match else store_code
            label_m = re.search(rf'{store_code}\s+-\s+(.+)', text)
            store_label = re.split(r'CONG TY TNHH THUC PHAM ANH HONG|CÔNG TY TNHH THỰC PHẨM ÁNH HỒNG',
                                    label_m.group(1))[0].strip() if label_m else None
            items = []
            for line in text.split("\n"):
                toks = line.split()
                if (len(toks) >= 6 and toks[0].isdigit() and toks[1].isdigit()
                        and re.match(r'^\d{13}$', toks[2])):
                    barcode = toks[2]
                    uom = toks[-4] if len(toks) >= 4 else None
                    try:
                        qty = int(float(toks[-2].replace(',', '')))
                    except ValueError:
                        continue
                    items.append((barcode, qty, uom))
            if items:
                orders.append({"vendor": "MINISTOP", "store_code": store_code,
                                "display_name": store_label,
                                "po_ref": po_ref, "items": items,
                                "page": page.page_number})
            else:
                print(f"  [CẢNH BÁO] {fname} trang {page.page_number} (CH {store_code}): "
                      f"nhận diện MINISTOP nhưng không đọc được dòng hàng nào -> BỎ QUA, cần kiểm tra tay")
    return orders


def parse_winmart(path):
    """Mỗi trang = 1 cửa hàng (1 file có thể chứa rất nhiều trang/đơn).
    KHÔNG đoán WinMart hay WinMart+ qua chữ viết tắt trên PO (WM/WM+/WIN/WIN+...
    vì cách viết không cố định và có thể xuất hiện thêm biến thể mới sau này).
    Chỉ lấy đúng MÃ SỐ cửa hàng; việc đây là WinMart hay WinMart+ giao hẳn cho
    MCL quyết định ở bước resolve_store (thử tra cả 2 hệ thống, mã CH của 2 bên
    không trùng nhau nên luôn ra đúng 1 kết quả)."""
    orders = []
    fname = os.path.basename(path)
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            if "WINCOMMERCE" not in text.upper():
                continue
            store_m = re.search(r'([0-9A-Z]{3,8})\s*-\s*(WM\+|WIN|WM)\s*(.+)', text)
            store_code = store_m.group(1) if store_m else None
            store_label = store_m.group(3).strip() if store_m else None
            vendor = "WINMART"  # nhãn chung, resolve_store sẽ thử cả WINMART/WINMART+
            po_m = re.search(r'PO No\.\)\s*(\d+)', text)
            po_ref = po_m.group(1) if po_m else None
            items = []
            for line in text.split("\n"):
                toks = line.split()
                if len(toks) >= 6 and toks[0].isdigit() and re.match(r'^\d{13}$', toks[2]):
                    try:
                        qty = int(round(float(toks[3].replace(',', '.'))))
                    except ValueError:
                        continue
                    items.append((toks[2], qty, None))
            if store_code and items:
                orders.append({"vendor": vendor, "store_code": store_code,
                                "display_name": store_label,
                                "po_ref": po_ref, "items": items,
                                "page": page.page_number})
            elif not store_code:
                print(f"  [CẢNH BÁO] {fname} trang {page.page_number}: nhận diện WinCommerce "
                      f"nhưng không đọc được mã cửa hàng -> BỎ QUA, cần kiểm tra tay")
            else:
                print(f"  [CẢNH BÁO] {fname} trang {page.page_number} (CH {store_code}): "
                      f"nhận diện WinCommerce nhưng không đọc được dòng hàng nào -> BỎ QUA, cần kiểm tra tay")
    return orders


def _undouble(s):
    """PDF CircleK in đậm phần header khiến mỗi ký tự bị lặp đôi. Gộp lại."""
    return ''.join(s[i] for i in range(0, len(s), 2))


def parse_circlek(path):
    orders = []
    fname = os.path.basename(path)
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            store_m = re.search(r'Store Code:\s*(\S+)', text)
            if not store_m:
                continue
            store_code = store_m.group(1)
            label_m = re.search(r'Store\(Warehouse\):\s*(.+?)\s*Người liên hệ', text, re.S)
            store_label = re.sub(r'\s+', ' ', label_m.group(1)).strip() if label_m else None
            header = text.split("STT")[0] if "STT" in text else text[:600]
            po_m = re.search(r'PR-[\w\-]+', _undouble(header))
            po_ref = po_m.group(0) if po_m else store_code
            items = []
            for line in text.split("\n"):
                m = re.match(
                    r'^\d+\s+\d+\s+(\d{13})\S*.*?\s+(HOP|HỘP|LOC|LỐC|LY|BICH|BỊCH|TÚI|TUI)\s+\d+\s+(\d+)\s+',
                    line, re.I)
                if m:
                    items.append((m.group(1), int(m.group(3)), m.group(2)))
            if items:
                orders.append({"vendor": "CIRCLEK", "store_code": store_code,
                                "display_name": store_label,
                                "po_ref": po_ref, "items": items,
                                "page": page.page_number})
            else:
                print(f"  [CẢNH BÁO] {fname} trang {page.page_number} (CH {store_code}): "
                      f"nhận diện CircleK nhưng không đọc được dòng hàng nào -> BỎ QUA, cần kiểm tra tay")
    return orders


_NAME_TO_BARCODE = {
    "thach dua anh hong bich 900g": "8935117702047",
    "thach dua anh hong bich 400g": "8935117702030",
    "thach dua ly anh hong 190g": "8935117702023",
    "banh flan caramel hop 80g": "8935117700128",
    "flan caramel cafe sua dua": "8935117700135",
}

# Co-opFood (JDA) hiện chỉ đặt hàng đúng 10 SKU cố định -> tra thẳng theo SKU
# (chính xác hơn nhiều so với đoán theo tên mô tả). Hậu tố sau dấu "-"
# (vd "3182634-4") là số check digit, không cố định theo sản phẩm nên bỏ qua.
COOPFOOD_SKU_TO_TP = {
    "3182634": "TP050005",  # Thach dua ANHHONG 6lyx190g       -> TDL 190g LỐC (đã xác nhận, không phải Ly)
    "3437206": "TP050007",  # Thach dua Anh Hong bich 400g     -> TDB 400g
    "3437211": "TP050006",  # Thach dua Anh Hong bich 900g     -> TDB 900g
    "3596799": "TP050037",  # Thach dua hdao A.HONG 6lyx190g   -> ĐÀO LỐC (đã xác nhận, không phải Ly)
    "3596800": "TP050039",  # Thachdua h.dua luoi AH6lyx190g   -> DƯA LƯỚI LỐC (đã xác nhận, không phải Ly)
    "3098253": "TP010004",  # Banh flan ANHHONG 2hopx100g      -> FĐ2 100g
    "3098254": "TP010008",  # Banh flan ANHHONG 10hopx54g      -> FL10 54g
    "3098275": "TP010007",  # Banh flan ANHHONG 6hopx54g       -> FL6 54g
    "3418233": "TP010014",  # B.Flan cao cap ANH HONG 2hx80g   -> FCC Lốc
    "3501144": "TP010016",  # B.Flan A.HONG cafesuadua 2x82g   -> CFSD Lốc
}


def _slug(s):
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode()
    return re.sub(r'\s+', ' ', s).strip().lower()


def _match_coopfood_product(desc_slug):
    """Suy ra mã TP trực tiếp từ mô tả sản phẩm (không dấu) trên PO Co-opFood/JDA.
    Bao phủ toàn bộ ~24 sản phẩm trong bảng giá, không chỉ vài sản phẩm cố định."""
    d = desc_slug
    if "cocochew" in d:
        if "chanh day" in d:
            return "TP050061"
        if "dua" in d:  # "dứa" (khóm) - không nhầm với "dừa" vì luôn đi kèm cocochew
            return "TP050062"
        return None
    if "caramel" in d:
        if "cf" in d or "ca phe" in d or "sua dua" in d:
            return "TP010016"  # CFSD Lốc
        return "TP010014"  # Flan Caramel cao cấp Lốc
    if "flan" in d:  # Flan truyền thống (không phải caramel)
        if "100" in d:
            return "TP010005" if "6" in d else "TP010004"
        if "54" in d:
            return "TP010008" if "10" in d else "TP010007"
        return None
    if "thach dua" in d:
        if "nho" in d and "90" in d:  # "nhỏ" 90g
            return "TP050002"
        if "dao" in d:  # vị đào
            return "TP050038" if "ly" in d else "TP050037"
        if "dua luoi" in d:  # dưa lưới
            return "TP050040" if "ly" in d else "TP050039"
        if "bich" in d:
            if "400" in d:
                return "TP050007"
            if "900" in d:
                return "TP050006"
            return None
        if "190" in d:
            return "TP050004" if "ly" in d else "TP050005"
        return None
    if "yuzu" in d and ("tui" in d or "goi" in d):
        if "500" in d:
            return "TP030039"
        if "1kg" in d or "1 kg" in d:
            return "TP030040"
        return None
    if "yuzu" in d and ("thung" in d or "6kg" in d):
        return "TP030019"
    if "dau" in d:
        return "TP020061"  # STC dâu
    if "nho" in d:
        return "TP020059"  # STC nho
    if "tao" in d:
        return "TP020057"  # STC táo
    if "cam" in d:
        return "TP020055"  # STC cam
    return None


def parse_coopfood(path):
    """Co-opFood / Co-opSmile / Co-opCheers dùng chung nền JDA, cùng bảng SKU.
    QUAN TRỌNG: mỗi TRANG là 1 P/O riêng biệt (khác P/O Number, khác P/O
    Location/Notes, khác cửa hàng) — 1 file có thể gộp nhiều chục trang/đơn,
    nên phải đọc RIÊNG TỪNG TRANG như MiniStop/WinMart/CircleK, KHÔNG được gộp
    cả file thành 1 khối text (làm vậy sẽ chỉ lấy được đúng 1 đơn, mất hết các
    đơn còn lại trong file mà không có cảnh báo gì).
    - Co-opFood gốc: 'Ship To' chỉ có mã kho (P/O Location) + tên viết tắt CH,
      không có địa chỉ -> tra MCL theo mã P/O Location như cũ.
    - Co-opSmile / Co-opCheers: 'Ship To' chỉ là mã kho vùng chung (không phải
      mã riêng từng CH) nhưng dòng Notes có tên hệ thống con + địa chỉ đầy đủ
      -> tra MCL theo địa chỉ (so khớp mờ), giống Satrafood/SaigonHD.
    """
    orders = []
    fname = os.path.basename(path)
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            if "JDA SOFTWARE" not in text.upper() and "P/O LOCATION" not in text.upper():
                continue
            po_m = re.search(r'P/O Number:\s*([\w]+)', text)  # bỏ phần "-00" phía sau
            po_ref = po_m.group(1) if po_m else None

            items = []
            for line in text.split("\n"):
                m = re.match(
                    r'^\s*(\d{6,8})-\d+\s+(.+?)\s+EA\s+\S+\s+[\d.]+\s+[\d.]+\s+([\d.]+)\s+([\d.]+)\s+[\d.]+\s+[\d,]+\.\d+\s*\*?\s*$',
                    line)
                if m:
                    sku = m.group(1)
                    desc = _slug(m.group(2))
                    qty = float(m.group(4))
                    # Ưu tiên tra theo SKU (chính xác) — chỉ fallback đoán theo tên
                    # nếu gặp SKU lạ chưa có trong danh mục 10 mã hiện tại.
                    tp = COOPFOOD_SKU_TO_TP.get(sku) or _match_coopfood_product(desc)
                    if tp:
                        items.append((tp, int(qty), None))  # đã là mã TP trực tiếp, không phải barcode
            if not items:
                print(f"  [CẢNH BÁO] {fname} trang {page.page_number}: nhận diện JDA/Co-opFood "
                      f"nhưng không đọc được dòng hàng nào -> BỎ QUA, cần kiểm tra tay")
                continue

            # Notes - CHEERS - RICHMOND CITY          * = This SKU Discounted
            #         207C Nguyen Xi, Phuong 26, Quan Binh Thanh
            notes_m = re.search(r'Notes\s*-\s*(CO\.?OPSMILE|CHEERS)\s*-\s*(.+?)\n\s*(.+?)\n', text, re.I)
            if notes_m:
                vendor = "COOPCHEERS" if "CHEERS" in notes_m.group(1).upper() else "COOPSMILE"
                ten_ch = re.sub(r'\*\s*=\s*This SKU Discounted', '', notes_m.group(2)).strip()
                diachi = notes_m.group(3).strip()
                orders.append({"vendor": vendor, "store_code": None,
                                "store_name": diachi, "store_name_alt": ten_ch,
                                "po_ref": po_ref, "items": items, "items_are_tp": True,
                                "page": page.page_number})
            else:
                loc_m = re.search(r'P/O Location:\s*(\S+)', text)
                store_code = loc_m.group(1) if loc_m else None
                ship_m = re.search(r'\d+-TC\s+CF\s+(.+?)\s+Contact', text)
                store_label = ship_m.group(1).strip() if ship_m else None
                if store_code:
                    orders.append({"vendor": "COOPFOOD", "store_code": store_code,
                                    "display_name": store_label,
                                    "po_ref": po_ref, "items": items, "items_are_tp": True,
                                    "page": page.page_number})
                else:
                    print(f"  [CẢNH BÁO] {fname} trang {page.page_number}: nhận diện Co-opFood "
                          f"nhưng không đọc được P/O Location -> BỎ QUA, cần kiểm tra tay")
    return orders


def parse_satrafood(path):
    """Không có mã CH số -> tra MCL theo (số nhà + tên đường) trích từ 'Nơi giao hàng'.
    Barcode đôi khi bị in dư 2 số ở đầu (vd '028935117700128' thay vì
    '8935117700128') -> luôn lấy 13 ký tự CUỐI của chuỗi số bắt được.
    Riêng mã Flan Caramel cao cấp (8935117700128): Satrafood luôn hiểu là Lốc
    dù PO ghi ĐVT là Hộp/Lốc2 -> không áp dụng đổi mã theo ĐVT (khác CircleK)."""
    orders = []
    with pdfplumber.open(path) as pdf:
        text = "\n".join((p.extract_text() or "") for p in pdf.pages)
    if "SATRAFOOD" not in text.upper():
        return orders
    po_m = re.search(r'(P-\d{6,})', text)
    po_ref = po_m.group(1) if po_m else None
    addr_m = re.search(r'Nơi giao hàng:\s*(.+)', text)
    diachi = addr_m.group(1).strip() if addr_m else None
    items = []
    for line in text.split("\n"):
        m = re.match(
            r'^\s*\d+\s+\d+\s+(\d{13,20})\s+.+?\s+(?:LỐC6|LỐC2|LỐC|BỊCH|HỘP)\s+([\d,]+)\s+',
            line, re.I)
        if m:
            barcode = m.group(1)[-13:]
            try:
                qty = int(round(float(m.group(2).replace(',', '.'))))
            except ValueError:
                continue
            items.append((barcode, qty, None))
    if diachi and items:
        orders.append({"vendor": "SATRAFOOD", "store_code": None, "store_name": diachi,
                        "po_ref": po_ref, "items": items})
    elif not diachi:
        print(f"  [CẢNH BÁO] {os.path.basename(path)}: nhận diện SATRAFOOD nhưng không đọc "
              f"được 'Nơi giao hàng' -> BỎ QUA, cần kiểm tra tay")
    elif not items:
        print(f"  [CẢNH BÁO] {os.path.basename(path)} ({diachi}): nhận diện SATRAFOOD nhưng "
              f"không đọc được dòng hàng nào -> BỎ QUA, cần kiểm tra tay")
    return orders


def parse_3sach(path):
    """Không có mã CH số -> tra MCL theo tên kho (WAREHOUSE/KHO) + địa chỉ.
    Mỗi dòng hàng: '<STT> <Mã 3S> <Tên hàng> <SL> <ĐVT> <Đơn giá> ...' rồi
    XUỐNG DÒNG mới tới barcode 13 số (barcode nằm dưới mã vạch ảnh, không cùng
    dòng với mô tả)."""
    orders = []
    with pdfplumber.open(path) as pdf:
        text = "\n".join((p.extract_text() or "") for p in pdf.pages)
    if "3SACH" not in text.upper():
        return orders
    po_m = re.search(r'\bPO(\d{6,})\b', text)
    po_ref = f"PO{po_m.group(1)}" if po_m else None
    wh_m = re.search(r'WAREHOUSE/\s*KHO:\s*(.+)', text)
    ten_kho = wh_m.group(1).strip() if wh_m else None
    addr_all = re.findall(r'Address/\s*Địa chỉ:\s*(.+)', text)
    diachi = addr_all[-1].strip() if addr_all else None

    lines = text.split("\n")
    items = []
    for i, line in enumerate(lines):
        m = re.match(
            r'^\s*(\d+)\s+(\d+)\s+(.+?)\s+([\d.]+)\s+(?:Ly|L[ốo]c|G[óo]i|B[ịi]ch|H[ộo]p)\s+',
            line, re.I)
        if m and i + 1 < len(lines):
            bc = lines[i + 1].strip()
            if re.match(r'^\d{13}$', bc):
                qty = int(round(float(m.group(4))))
                items.append((bc, qty, None))
    if (ten_kho or diachi) and items:
        orders.append({"vendor": "3SACH", "store_code": None,
                        "store_name": ten_kho, "store_name_alt": diachi,
                        "po_ref": po_ref, "items": items})
    elif not (ten_kho or diachi):
        print(f"  [CẢNH BÁO] {os.path.basename(path)}: nhận diện 3SACH nhưng không đọc được "
              f"WAREHOUSE/địa chỉ -> BỎ QUA, cần kiểm tra tay")
    elif not items:
        print(f"  [CẢNH BÁO] {os.path.basename(path)} ({ten_kho}): nhận diện 3SACH nhưng "
              f"không đọc được dòng hàng nào -> BỎ QUA, cần kiểm tra tay")
    return orders


def parse_siba(path):
    """Có mã CH (VD 'S102') trong 'Nơi nhận hàng' -> tra trực tiếp theo mã CH;
    nếu MCL chưa kịp cập nhật mã CH này thì fallback fuzzy-match theo địa chỉ.
    Bảng hàng có viền rõ ràng -> dùng pdfplumber.extract_tables() thay vì regex
    trên text thô (ô 'Mô tả MH' lẫn nhiều số liệu gây nhiễu nếu đọc theo dòng)."""
    orders = []
    with pdfplumber.open(path) as pdf:
        text = "\n".join((p.extract_text() or "") for p in pdf.pages)
        if "SIBA FOOD" not in text.upper():
            return orders
        po_m = re.search(r'Số đơn hàng:\s*(\S+)', text)
        po_ref = po_m.group(1) if po_m else None
        store_m = re.search(r'Nơi nhận hàng:\s*(\S+)\s*-\s*(.+?)\n\s*Diễn giải', text, re.S)
        store_code = store_m.group(1) if store_m else None
        diachi = re.sub(r'\s+', ' ', store_m.group(2)).strip() if store_m else None

        items = []
        for page in pdf.pages:
            for table in (page.extract_tables() or []):
                if not table:
                    continue
                header = [(c or "").replace("\n", " ").strip().lower() for c in table[0]]
                try:
                    idx_bc = next(i for i, h in enumerate(header) if "vạch" in h)
                    idx_qty = next(i for i, h in enumerate(header) if "lượng" in h)
                except StopIteration:
                    continue
                for row in table[1:]:
                    if idx_bc >= len(row) or idx_qty >= len(row):
                        continue
                    bc = re.sub(r'\s+', '', row[idx_bc] or '')
                    if not re.match(r'^\d{13}$', bc):
                        continue
                    qty_cell = (row[idx_qty] or "").strip()
                    try:
                        qty = int(round(float(qty_cell.replace(',', '.'))))
                    except ValueError:
                        continue
                    items.append((bc, qty, None))
    if (store_code or diachi) and items:
        orders.append({"vendor": "SIBA", "store_code": store_code, "store_name": diachi,
                        "po_ref": po_ref, "items": items})
    elif not (store_code or diachi):
        print(f"  [CẢNH BÁO] {os.path.basename(path)}: nhận diện SIBA nhưng không đọc được "
              f"'Nơi nhận hàng' -> BỎ QUA, cần kiểm tra tay")
    elif not items:
        print(f"  [CẢNH BÁO] {os.path.basename(path)} (CH {store_code}): nhận diện SIBA nhưng "
              f"không đọc được dòng hàng nào -> BỎ QUA, cần kiểm tra tay")
    return orders


def parse_menas(path):
    """Không có mã CH số -> tra MCL theo tên siêu thị + địa chỉ giao hàng
    (SHIP ADDRESS), giống cơ chế Satrafood/3Sach."""
    orders = []
    with pdfplumber.open(path) as pdf:
        text = "\n".join((p.extract_text() or "") for p in pdf.pages)
    if "MENAS" not in text.upper():
        return orders
    po_m = re.search(r'\(SHIP ADDRESS\)\s*(\S+)', text)
    po_ref = po_m.group(1) if po_m else None
    ten_m = re.search(r'ĐỊA CHỈ GIAO HÀNG.*?\n(.+?)\s+NGÀY ĐẶT HÀNG', text, re.S)
    ten_ch = ten_m.group(1).strip() if ten_m else None
    addr_m = re.search(r'\n(\d+\s+\S.+?)\s*NGÀY GIAO HÀNG', text, re.S)
    diachi = re.sub(r'\s+', ' ', addr_m.group(1)).strip() if addr_m else None

    items = []
    for line in text.split("\n"):
        m = re.match(r'^\d+\s+M\d+\s+(\d{13})\s+.+?\s+LOC\s+[\d.]+\s+(\d+)\s+', line)
        if m:
            items.append((m.group(1), int(m.group(2)), None))

    if (ten_ch or diachi) and items:
        orders.append({"vendor": "MENAS", "store_code": None,
                        "store_name": diachi, "store_name_alt": ten_ch,
                        "po_ref": po_ref, "items": items})
    elif not (ten_ch or diachi):
        print(f"  [CẢNH BÁO] {os.path.basename(path)}: nhận diện MENAS nhưng không đọc được "
              f"'ĐỊA CHỈ GIAO HÀNG' -> BỎ QUA, cần kiểm tra tay")
    elif not items:
        print(f"  [CẢNH BÁO] {os.path.basename(path)} ({ten_ch}): nhận diện MENAS nhưng "
              f"không đọc được dòng hàng nào -> BỎ QUA, cần kiểm tra tay")
    return orders


def parse_saigonhd(path):
    """Không có mã cửa hàng số -> tra MCL theo ĐỊA CHỈ (so khớp mờ)."""
    orders = []
    with pdfplumber.open(path) as pdf:
        text = "\n".join((p.extract_text() or "") for p in pdf.pages)
    addr_m = re.search(r'Đơn vị đặt hàng:\s*(.+)', text)
    diachi = addr_m.group(1).strip() if addr_m else None
    po_m = re.search(r'Số đơn hàng:\s*(\S+)', text)
    po_ref = po_m.group(1) if po_m else None
    items = []
    for line in text.split("\n"):
        toks = line.split()
        if len(toks) >= 12 and toks[0].isdigit() and re.match(r'^\d{13}$', toks[2]):
            unit, qty_str = toks[-10], toks[-9]
            if not re.match(r'^[\d,]+$', qty_str):
                continue
            try:
                qty = int(round(float(qty_str.replace(',', '.'))))
            except ValueError:
                continue
            items.append((toks[2], qty, unit))
    if diachi and items:
        orders.append({"vendor": "SÀIGÒNHD", "store_code": None, "store_name": diachi,
                        "po_ref": po_ref, "items": items})
    elif not diachi:
        print(f"  [CẢNH BÁO] {os.path.basename(path)}: nhận diện SaigonHD nhưng không đọc "
              f"được 'Đơn vị đặt hàng' -> BỎ QUA, cần kiểm tra tay")
    elif not items:
        print(f"  [CẢNH BÁO] {os.path.basename(path)} ({diachi}): nhận diện SaigonHD nhưng "
              f"không đọc được dòng hàng nào -> BỎ QUA, cần kiểm tra tay")
    return orders


_OSI_ITEM_RE = re.compile(
    r'^\d+\s+(\d{13})\s+\S+\s+.*?\s+[\d.]+\s+[\d,]+\s+\S+\s+\S+\s+([\d.]+)\s+[\d.]+\s+[\d,]+')


def parse_osifood(path):
    """Logo 'OsiFood' là ảnh (không có trong text) -> nhận diện qua tiêu đề cột.
    Không có mã cửa hàng khớp MCL -> tra theo ĐỊA CHỈ (so khớp mờ)."""
    orders = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            if "MÃ GỢI NHỚ" not in text.upper():
                continue
            addr_m = re.search(r'ĐỊA CHỈ\s+(.+?)\n\s*GIAO HÀNG\s+(.+)', text)
            diachi = f"{addr_m.group(1)} {addr_m.group(2)}".strip() if addr_m else None
            ten_m = re.search(r'TÊN CH\s+(.+?)\s+Ngày đặt hàng', text)
            ten_ch = ten_m.group(1).strip() if ten_m else None
            store_m = re.search(r'STORE\s+(\d+)', text)
            store_code = store_m.group(1) if store_m else None
            date_m = re.search(r'Ngày đặt hàng\s*:\s*(\S+)', text)
            po_ref = f"OSI-{store_code}-{date_m.group(1)}" if store_code and date_m else store_code
            items = []
            for line in text.split("\n"):
                m = _OSI_ITEM_RE.match(line)
                if m:
                    items.append((m.group(1), int(round(float(m.group(2)))), None))
            if (diachi or ten_ch) and items:
                orders.append({"vendor": "OSIFOOD", "store_code": store_code,
                                "store_name": diachi, "store_name_alt": ten_ch,
                                "po_ref": po_ref, "items": items,
                                "page": page.page_number})
            elif not (diachi or ten_ch):
                print(f"  [CẢNH BÁO] {os.path.basename(path)} trang {page.page_number}: "
                      f"nhận diện OsiFood nhưng không đọc được địa chỉ/tên CH -> BỎ QUA, cần kiểm tra tay")
            elif not items:
                print(f"  [CẢNH BÁO] {os.path.basename(path)} trang {page.page_number} "
                      f"({diachi or ten_ch}): nhận diện OsiFood nhưng không đọc được dòng hàng nào "
                      f"-> BỎ QUA, cần kiểm tra tay")
    return orders


VENDOR_DETECT = [
    ("MINISTOP", lambda t: "MINISTOP" in t.upper(), parse_ministop),
    ("WINMART", lambda t: "WINCOMMERCE" in t.upper(), parse_winmart),
    ("CIRCLEK", lambda t: "VÒNG TRÒN ĐỎ" in t.upper() or "CIRCLE K" in t.upper() or "CIRCLEK" in t.upper(), parse_circlek),
    ("SATRAFOOD", lambda t: "SATRAFOOD" in t.upper(), parse_satrafood),
    ("3SACH", lambda t: "3SACH" in t.upper(), parse_3sach),
    ("SIBA", lambda t: "SIBA FOOD" in t.upper(), parse_siba),
    ("MENAS", lambda t: "MENAS" in t.upper(), parse_menas),
    ("COOPFOOD", lambda t: "JDA SOFTWARE" in t.upper() or "P/O LOCATION" in t.upper(), parse_coopfood),
    ("SÀIGÒNHD", lambda t: "SÀI GÒN HD" in t.upper() or "SAI GON HD" in t.upper(), parse_saigonhd),
    ("OSIFOOD", lambda t: "MÃ GỢI NHỚ" in t.upper(), parse_osifood),
]


def detect_and_parse(pdf_path):
    try:
        with pdfplumber.open(pdf_path) as pdf:
            first_text = pdf.pages[0].extract_text() or ""
    except Exception as e:
        print(f"  [LỖI] Không đọc được {pdf_path}: {e}")
        return []
    for name, matcher, fn in VENDOR_DETECT:
        if matcher(first_text):
            try:
                return fn(pdf_path)
            except Exception as e:
                print(f"  [LỖI] Parser {name} lỗi trên {pdf_path}: {e}")
                return []
    print(f"  [BỎ QUA] Không nhận diện được hệ thống: {os.path.basename(pdf_path)}")
    return []


# =============================================================================
# 3. TRA CỨU MCL
#    - Đa số hệ thống: tra theo MÃ CH THEO HỆ THỐNG (chính xác).
#    - Sài Gòn HD / OsiFood: cột mã bị trống trong MCL -> so khớp mờ theo địa chỉ/tên.
# =============================================================================

def build_mcl(template_path):
    wb = openpyxl.load_workbook(template_path, data_only=True)
    ws = wb['MCL']
    code_lookup = {}
    name_pool = {}  # hệ thống -> list[(mã ERP, tên siêu thị normalize)]
    for r in range(4, ws.max_row + 1):
        ma_erp = ws.cell(row=r, column=3).value
        ten = ws.cell(row=r, column=4).value
        he_thong = ws.cell(row=r, column=9).value
        ma_ch = ws.cell(row=r, column=10).value
        if not (ma_erp and he_thong):
            continue
        ht = str(he_thong).strip().upper()
        if ma_ch is not None:
            code_lookup[f"{ht}||{str(ma_ch).strip()}"] = ma_erp
        if ten:
            name_pool.setdefault(ht, []).append((ma_erp, _normalize(str(ten))))
    return code_lookup, name_pool


_VENDOR_STRIP_WORDS = {
    "OSIFOOD": ["OSIFOOD"],
    "SÀIGÒNHD": ["SAIGONHD", "SAI GON HD", "CONG TY CO PHAN SAI GON HD"],
    "SATRAFOOD": ["SATRAFOODS", "SATRAFOOD"],
    "COOPSMILE": ["CO OPSMILE", "COOPSMILE", "CO SMILE"],
    "COOPCHEERS": ["CHEERS"],
    "3SACH": ["3SACH", "3 SACH"],
    "SIBA": ["SIBA FOOD", "SIBA"],
    "MENAS": ["MENAS"],
}


# Tiền tố loại hình doanh nghiệp xuất hiện phổ biến trong tên đầy đủ ở MCL
# (VD "CÔNG TY CỔ PHẦN SÀI GÒN HD (PEARL)") -> luôn xoá trước tiên, áp dụng
# cho MỌI hệ thống, không riêng gì SÀIGÒNHD.
_GENERIC_COMPANY_WORDS = ["CONG TY CO PHAN", "CONG TY TNHH MTV", "CONG TY TNHH", "CTY CP", "CTY"]


def _strip_vendor_words(s, vendor):
    # Xoá cụm DÀI trước, cụm NGẮN sau — tránh trường hợp cụm ngắn "cắn" mất 1
    # phần của cụm dài hơn khiến cụm dài không còn khớp được nữa (từng xảy ra
    # với SÀIGÒNHD: "SAI GON HD" xoá trước làm hỏng "CONG TY CO PHAN SAI GON HD").
    words = _GENERIC_COMPANY_WORDS + _VENDOR_STRIP_WORDS.get(vendor, [])
    for w in sorted(set(words), key=len, reverse=True):
        s = s.replace(w, " ")
    return re.sub(r'\s+', ' ', s).strip()


# Các nhóm hệ thống dùng chung 1 định dạng PDF nhưng MCL lại tách hệ thống
# riêng theo mã CH (mã CH của các hệ thống con trong 1 nhóm không trùng nhau)
# -> khi tra theo mã, thử lần lượt từng hệ thống trong nhóm, không cần đoán
# qua chữ viết tắt trên PO (vốn không ổn định, dễ có biến thể mới).
_VENDOR_CODE_GROUPS = {
    "WINMART": ["WINMART", "WINMART+"],
}


# Từ đệm địa chỉ xuất hiện ở HẦU HẾT mọi địa chỉ (đường, số, phường, quận,
# thành phố...) -> không có giá trị phân biệt cửa hàng này với cửa hàng khác,
# ngược lại còn gây khớp nhầm ngẫu nhiên (VD "48 Duong Thi Muoi" và "416 Duong
# Quang Ham" cùng chứa "DUONG" nên bị đẩy điểm khớp lên dù là 2 địa chỉ khác
# hẳn nhau) -> lọc bỏ trước khi so điểm.
_ADDRESS_FILLER_WORDS = {
    "DUONG", "SO", "PHUONG", "P", "QUAN", "Q", "KHU", "PHO", "TP", "THANH",
    "VIET", "NAM", "HO", "CHI", "MINH", "KP", "TO", "AP", "TINH", "HUYEN",
}


def _strip_address_filler(s):
    return " ".join(w for w in s.split() if w not in _ADDRESS_FILLER_WORDS)


def resolve_store(order, code_lookup, name_pool, fuzzy_threshold=0.35, confident_threshold=0.6):
    vendor = order["vendor"]
    if order.get("store_code") is not None:
        for v in _VENDOR_CODE_GROUPS.get(vendor, [vendor]):
            key = f"{v}||{order['store_code']}"
            if key in code_lookup:
                return code_lookup[key], "code", None
    candidates = [n for n in (order.get("store_name"), order.get("store_name_alt")) if n]
    if candidates and vendor in name_pool:
        best_score, best_erp, best_candidate_raw = 0.0, None, None
        for target in candidates:
            norm_target = _strip_address_filler(_normalize(target))
            for ma_erp, norm_name in name_pool[vendor]:
                candidate = _strip_address_filler(_strip_vendor_words(norm_name, vendor))
                score = difflib.SequenceMatcher(None, norm_target, candidate).ratio()
                if candidate and candidate in norm_target:
                    score = max(score, 0.6)
                if score > best_score:
                    best_score, best_erp, best_candidate_raw = score, ma_erp, norm_name
        if best_score >= confident_threshold:
            return best_erp, f"fuzzy({best_score:.2f})", None
        if best_score >= fuzzy_threshold:
            warn = f"KIỂM TRA: khớp mờ \"{best_candidate_raw}\" ({best_erp}, điểm {best_score:.2f})"
            return best_erp, f"fuzzy_yếu({best_score:.2f})", warn
    return None, None, None


# =============================================================================
# 4. GHÉP TẤT CẢ LẠI VÀ GHI VÀO FILE EXCEL
# =============================================================================

class _WarningCounter:
    """'Nghe lén' stdout trong lúc đọc PDF để đếm tổng số dòng [CẢNH BÁO] đã in ra,
    dùng cho dòng tổng kết cuối cùng — không cần sửa từng chỗ print() cảnh báo."""
    def __init__(self, real):
        self.real = real
        self.count = 0

    def write(self, s):
        self.real.write(s)
        self.count += s.count("[CẢNH BÁO]")

    def flush(self):
        self.real.flush()


# Thứ tự ưu tiên khi ghi vào form (theo yêu cầu anh Long) — trong cùng 1 hệ
# thống vẫn giữ nguyên thứ tự đọc file/PDF, không sắp xếp thêm.
VENDOR_ORDER = ["MINISTOP", "WINMART", "WINMART+", "COOPFOOD", "COOPSMILE",
                "COOPCHEERS", "SATRAFOOD", "SÀIGÒNHD", "OSIFOOD", "3SACH",
                "SIBA", "MENAS", "CIRCLEK"]


def process_folder(folder):
    pdf_files = sorted(set(glob.glob(os.path.join(folder, "*.pdf")) +
                            glob.glob(os.path.join(folder, "*.PDF"))))
    xlsx_files = [f for f in glob.glob(os.path.join(folder, "*.xlsx"))
                  if "đã_điền" not in f and "da_dien" not in f]
    if not xlsx_files:
        print("KHÔNG tìm thấy file Form_Hàng_MT10.xlsx trong thư mục.")
        sys.exit(1)
    template_path = xlsx_files[0]
    print(f"Dùng form mẫu: {os.path.basename(template_path)}")
    print(f"Tìm thấy {len(pdf_files)} file PDF PO.\n")

    code_lookup, name_pool = build_mcl(template_path)

    all_orders = []
    page_count_by_file = {}
    tee = _WarningCounter(sys.stdout)
    old_stdout = sys.stdout
    sys.stdout = tee
    try:
        for pdf_path in pdf_files:
            print(f"Đang đọc: {os.path.basename(pdf_path)}")
            orders = detect_and_parse(pdf_path)
            print(f"  -> nhận diện được {len(orders)} đơn hàng")
            for o in orders:
                o["src_file"] = pdf_path
            all_orders.extend(orders)
            try:
                with pdfplumber.open(pdf_path) as pdf:
                    page_count_by_file[pdf_path] = len(pdf.pages)
            except Exception:
                page_count_by_file[pdf_path] = 1
    finally:
        sys.stdout = old_stdout
    warning_count = tee.count

    rows = []
    resolved_pages_by_file = {}
    unresolved_stores, unresolved_products, weak_matches = [], [], []
    for o in all_orders:
        ma_erp, how, weak_warning = resolve_store(o, code_lookup, name_pool)
        if not ma_erp:
            label = o.get("store_code") or o.get("store_name") or "?"
            ten = o.get("display_name") or o.get("store_name_alt")
            suffix = f" ({ten})" if ten and ten != label else ""
            unresolved_stores.append(f"{o['vendor']}||{label}{suffix}")
            continue
        resolved_pages_by_file.setdefault(o["src_file"], []).append(o.get("page", 1))
        row = {"ma_erp": ma_erp, "vendor": o["vendor"], "po_ref": o["po_ref"],
               "qty_by_col": {}, "weak_warning": weak_warning}
        if weak_warning:
            label = o.get("store_code") or o.get("store_name") or "?"
            weak_matches.append(f"{o['vendor']}||{label}: {weak_warning} (PO {o.get('po_ref')})")
        for barcode_or_tp, qty, uom in o["items"]:
            if o.get("items_are_tp"):
                tp, adj_qty = barcode_or_tp, qty
            else:
                tp, adj_qty = resolve_tp(barcode_or_tp, uom, vendor=o["vendor"], qty=qty)
            if not tp or tp not in TP_TO_COL:
                unresolved_products.append((o["vendor"], o.get("store_code") or o.get("store_name"), barcode_or_tp, uom))
                continue
            col = TP_TO_COL[tp]
            row["qty_by_col"][col] = row["qty_by_col"].get(col, 0) + adj_qty
        rows.append(row)

    # Sắp theo thứ tự hệ thống đã chốt; sort() ổn định nên các dòng cùng hệ
    # thống vẫn giữ nguyên thứ tự đọc file như cũ.
    rows.sort(key=lambda r: VENDOR_ORDER.index(r["vendor"]) if r["vendor"] in VENDOR_ORDER
              else len(VENDOR_ORDER))

    print(f"\n{'='*60}")
    print(f"Tổng số file PDF đã đọc : {len(pdf_files)}")
    print(f"Tổng đơn parse được     : {len(all_orders)}")
    print(f"Khớp được với MCL       : {len(rows)}")
    print(f"KHÔNG khớp MCL          : {len(set(unresolved_stores))} mã/tên cửa hàng")
    print(f"Trang/file BỊ SÓT       : {warning_count} "
          f"(xem chi tiết [CẢNH BÁO] phía trên — đây là các trang nhận diện được "
          f"hệ thống nhưng KHÔNG tạo ra đơn hàng, cần đối chiếu tay)")
    if unresolved_stores:
        print("  Danh sách chưa khớp (cần bổ sung MCL hoặc kiểm tra tay):")
        for s in sorted(set(unresolved_stores)):
            print(f"    - {s}")
    multi_page_files = [f for f in pdf_files if page_count_by_file.get(f, 1) >= 2]
    if multi_page_files:
        print("\nCác trang đã lên form (để in), chỉ file từ 2 trang trở lên:")
        for f in multi_page_files:
            pages = sorted(set(resolved_pages_by_file.get(f, [])))
            pages_str = ",".join(str(x) for x in pages) if pages else "(không có trang nào khớp)"
            print(f"  {os.path.basename(f)}: {pages_str}")

    if weak_matches:
        print(f"\n[CẢNH BÁO] {len(weak_matches)} dòng khớp MỜ (đã ghi vào form, cột F có "
              f"ghi chú cảnh báo, cần soát lại tay):")
        for w in weak_matches:
            print(f"    - {w}")

    if unresolved_products:
        print(f"KHÔNG nhận diện được sản phẩm ({len(unresolved_products)} dòng):")
        for p in unresolved_products:
            print(f"    - {p}")

    out_path = os.path.join(folder, "Form_Hàng_MT10_đã_điền.xlsx")
    shutil.copy(template_path, out_path)
    wb = openpyxl.load_workbook(out_path, data_only=False)
    ws = wb['BÁN']

    # Dọn sạch một phạm vi RỘNG trước khi ghi (không chỉ 19 dòng) để không sót
    # dữ liệu cũ khi có nhiều đơn (thực tế có ngày 50-70 đơn).
    start_row = 24
    clear_until = max(345, start_row + len(rows) + 20)
    for r in range(start_row, clear_until + 1):
        for c in range(1, 39):
            ws.cell(row=r, column=c).value = None

    for i, row in enumerate(rows):
        r = start_row + i
        ws.cell(row=r, column=1).value = f'=IF(D{r}<>"",SUBTOTAL(103,$D$23:D{r}),"")'
        ws.cell(row=r, column=2).value = f'=IFERROR(VLOOKUP(C{r},MCL!$C$4:$J$1197,8,0),"")'
        ws.cell(row=r, column=3).value = row['ma_erp']
        ws.cell(row=r, column=4).value = f'=IFERROR(VLOOKUP($C{r},MCL!$C$4:$D$1659,2,0),"")'
        ws.cell(row=r, column=5).value = f'=IFERROR(VLOOKUP(C{r},MCL!$C$4:$H$650,6,0),"")'
        # Cột F (SO): để trống như thường lệ để anh dán tay số SO; NHƯNG nếu
        # dòng này khớp cửa hàng theo kiểu "khớp mờ yếu" (độ tin cậy thấp,
        # nghi ngờ sai) thì ghi chữ cảnh báo vào đây để anh thấy ngay trên
        # form, thay vì phải lật console. Khi anh dán số SO thật vào, chữ
        # cảnh báo tự bị ghi đè, không ảnh hưởng gì.
        if row.get('weak_warning'):
            ws.cell(row=r, column=6).value = row['weak_warning']
        for col_letter, qty in row['qty_by_col'].items():
            ws[f'{col_letter}{r}'] = qty
        ws.cell(row=r, column=34).value = f'=IFERROR(VLOOKUP($C{r},MCL!$C$4:$I$659,7,0),"")'
        ws.cell(row=r, column=35).value = ArrayFormula(
            f"AI{r}",
            f'=IFERROR(SUMPRODUCT($G{r}:$AG{r},INDEX($G$6:$AG$19,MATCH(IF($AH{r}="COOP","COOPFOOD",$AH{r}),$F$6:$F$19,0),0)),"")')
        ws.cell(row=r, column=36).value = ArrayFormula(
            f"AJ{r}",
            f'=IF($AH{r}="COOPFOOD",SUMPRODUCT($G{r}:$AG{r},$G$7:$AG$7),SUMPRODUCT($G{r}:$AG{r},$G$19:$AG$19))')
        ws.cell(row=r, column=37).value = f'=IFERROR(AI{r}*1.08,"")'
        ws.cell(row=r, column=38).value = row['po_ref']  # AL — số PO gốc

    ws['D2'] = datetime.datetime.now()
    wb.save(out_path)
    print(f"\n✅ Đã lưu kết quả: {out_path}")
    return out_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Dùng: python3 process_po.py \"/đường/dẫn/tới/thư_mục\"")
        sys.exit(1)
    process_folder(sys.argv[1])
